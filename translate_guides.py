#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Preparazione controllata delle localizzazioni documentali di ATK-Pro.

Senza ``--execute`` lo script svolge soltanto un audit: non carica chiavi API,
non effettua chiamate esterne e non scrive file. Con ``--execute`` genera i
documenti in ``localization_staging``; non modifica mai ``assets``.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import html as html_module
import json
import os
from pathlib import Path
import re
import sys
import time
from typing import Any
import urllib.error
import urllib.request


if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")


PROJECT_ROOT = Path(__file__).resolve().parent
ASSETS_DIR = PROJECT_ROOT / "assets"
ITALIAN_DIR = ASSETS_DIR / "it" / "testuali"
DEFAULT_STAGING_ROOT = PROJECT_ROOT / "localization_staging"
CONFIG_FILE = PROJECT_ROOT / "config.json"
DEFAULT_TRANSLATION_MEMORY = PROJECT_ROOT / "scripts" / "smartcat_translation_template.xlsx"

GUIDE_FILES = (
    "guida.html",
    "guida_01_installazione_configurazione.html",
    "guida_02_operazioni_base.html",
    "guida_03_ricerca_assistita_ai.html",
    "guida_04_visualizzazione_immagini.html",
    "guida_05_visualizzazione_metadati.html",
    "guida_06_ocr_avanzato.html",
    "guida_07_traduzione.html",
    "guida_08_esportazione_gedcom.html",
    "guida_09_supporto_faq.html",
)

PRESENTATION_FILES = (
    "presentazione_autore.html",
    "presentazione_progetto_ATK-Pro.html",
)

DISCLAIMER_FILES = (
    "disclaimer_legale_ATK-Pro.html",
    "disclaimer_legale_ATK-Pro.txt",
)

DOCUMENT_FILES = GUIDE_FILES + PRESENTATION_FILES + DISCLAIMER_FILES

# codice: (nome usato nel prompt, attributo HTML lang, direzione)
LANGUAGES = {
    "ar": ("Arabic", "ar", "rtl"),
    "da": ("Danish", "da", "ltr"),
    "de": ("German", "de", "ltr"),
    "el": ("Greek", "el", "ltr"),
    "en": ("English", "en", "ltr"),
    "es": ("Spanish", "es", "ltr"),
    "fr": ("French", "fr", "ltr"),
    "he": ("Hebrew", "he", "rtl"),
    "ja": ("Japanese", "ja", "ltr"),
    "nl": ("Dutch", "nl", "ltr"),
    "no": ("Norwegian", "no", "ltr"),
    "pl": ("Polish", "pl", "ltr"),
    "pt": ("Portuguese", "pt", "ltr"),
    "ro": ("Romanian", "ro", "ltr"),
    "ru": ("Russian", "ru", "ltr"),
    "sv": ("Swedish", "sv", "ltr"),
    "tr": ("Turkish", "tr", "ltr"),
    "vi": ("Vietnamese", "vi", "ltr"),
    "zh": ("Chinese (Simplified)", "zh", "ltr"),
}

MEMORY_HEADERS = {
    "ar": "إيطالي", "da": "Dansk", "de": "Deutsch", "el": "Ελληνικά",
    "en": "English", "es": "Español", "fr": "Français", "he": "איטלקית",
    "ja": "日本語", "nl": "Nederlands", "no": "Norsk", "pl": "Polski",
    "pt": "Português", "ro": "Română", "ru": "Русский", "sv": "Svenska",
    "tr": "Türkçe", "vi": "Tiếng Việt", "zh": "中文",
}

DO_NOT_TRANSLATE = (
    "ATK-Pro",
    "Antenati Toolkit Pro",
    "Portale Antenati",
    "GNU Affero General Public License",
    "AGPL-3.0-or-later",
    "Creative Commons Attribution-NonCommercial-ShareAlike 4.0 International",
    "CC BY-NC-SA 4.0",
    "GEDCOM",
    "TEI",
    "IIIF",
    "OCR",
    "LLM",
    "Gemini",
    "GPT-4o",
    "Claude",
    "Google",
    "OpenAI",
    "Anthropic",
    "XML",
    "JSON",
    "HTML",
    "DOCX",
    "PDF",
    "PNG",
    "JPG",
    "TIFF",
    "WebP",
    "EXIF",
    "API",
)

TRANSLATE_TAGS = {
    "title", "h1", "h2", "h3", "h4", "p", "li", "td", "th", "span",
    "a", "strong", "em", "button", "label", "caption", "dt", "dd",
    "figcaption", "blockquote", "pre",
}
SKIP_TAGS = {"style", "script", "code"}
PLACEHOLDER_RE = re.compile(r"\{(T\d{5})\}")
LINK_RE = re.compile(r'href=["\']([^"\']+\.html)["\']', re.IGNORECASE)


def sha256_text(text: str) -> str:
    return hashlib.sha256(text.encode("utf-8")).hexdigest()


def should_translate(text: str) -> bool:
    value = text.strip()
    if not value or re.fullmatch(r"[\s\d\W]+", value):
        return False
    if re.match(r"https?://", value) or re.match(r"^mailto:", value):
        return False
    if re.fullmatch(r"[\w.-]+@[\w.-]+", value):
        return False
    if re.fullmatch(r"[a-z0-9_.-]+\.(?:html|txt|webp|png|jpg|css|js)", value, re.I):
        return False
    return bool(re.search(r"[A-Za-zÀ-ÿ]", value))


def extract_html_text(html: str) -> tuple[str, dict[str, str]]:
    """Sostituisce nodi testuali e ``alt`` traducibili con segnaposto."""
    parts = re.split(r"(<[^>]*?>)", html, flags=re.DOTALL)
    texts: dict[str, str] = {}
    stack: list[str] = []
    skip_depth = 0
    result: list[str] = []

    for index, segment in enumerate(parts):
        if index % 2:
            def replace_attribute(match: re.Match[str]) -> str:
                value = match.group(3)
                if not should_translate(value):
                    return match.group(0)
                placeholder = f"T{len(texts):05d}"
                texts[placeholder] = value
                return f"{match.group(1)}{match.group(2)}{{{placeholder}}}{match.group(2)}"

            segment = re.sub(
                r"(\balt\s*=\s*)([\"'])(.*?)\2",
                replace_attribute,
                segment,
                flags=re.IGNORECASE | re.DOTALL,
            )
            if re.match(r"<meta\b", segment, flags=re.IGNORECASE) and re.search(
                r"\bname\s*=\s*([\"'])description\1", segment, flags=re.IGNORECASE,
            ):
                segment = re.sub(
                    r"(\bcontent\s*=\s*)([\"'])(.*?)\2",
                    replace_attribute,
                    segment,
                    count=1,
                    flags=re.IGNORECASE | re.DOTALL,
                )
            match = re.match(r"</?([a-zA-Z][a-zA-Z0-9]*)", segment)
            if match:
                tag = match.group(1).lower()
                closing = segment.startswith("</")
                self_closing = segment.rstrip().endswith("/>")
                if closing:
                    if tag in SKIP_TAGS and skip_depth:
                        skip_depth -= 1
                    for position in range(len(stack) - 1, -1, -1):
                        if stack[position] == tag:
                            stack.pop(position)
                            break
                elif not self_closing:
                    if tag in SKIP_TAGS:
                        skip_depth += 1
                    stack.append(tag)
            result.append(segment)
            continue

        stripped = segment.strip()
        if skip_depth or not any(tag in TRANSLATE_TAGS for tag in stack) or not should_translate(stripped):
            result.append(segment)
            continue
        placeholder = f"T{len(texts):05d}"
        texts[placeholder] = stripped
        leading = segment[: len(segment) - len(segment.lstrip())]
        trailing = segment[len(segment.rstrip()) :]
        result.append(f"{leading}{{{placeholder}}}{trailing}")

    return "".join(result), texts


def extract_plain_text(text: str) -> tuple[str, dict[str, str]]:
    """Tratta ogni riga non vuota del disclaimer TXT come unità controllata."""
    texts: dict[str, str] = {}
    result: list[str] = []
    for line in text.splitlines(keepends=True):
        ending = "\r\n" if line.endswith("\r\n") else "\n" if line.endswith("\n") else ""
        value = line[: -len(ending)] if ending else line
        if should_translate(value):
            placeholder = f"T{len(texts):05d}"
            texts[placeholder] = value
            result.append(f"{{{placeholder}}}{ending}")
        else:
            result.append(line)
    return "".join(result), texts


def apply_translations(template: str, source: dict[str, str], translated: dict[str, Any]) -> str:
    if set(translated) != set(source):
        missing = sorted(set(source) - set(translated))
        extra = sorted(set(translated) - set(source))
        raise ValueError(f"Segnaposto non allineati; mancanti={missing}, extra={extra}")
    if not all(isinstance(value, str) and value.strip() for value in translated.values()):
        raise ValueError("La risposta contiene traduzioni vuote o non testuali")
    rendered = PLACEHOLDER_RE.sub(lambda match: translated[match.group(1)], template)
    if PLACEHOLDER_RE.search(rendered):
        raise ValueError("Sono rimasti segnaposto non sostituiti")
    return rendered


def recover_translations(
    template: str, source: dict[str, str], rendered: str,
) -> dict[str, str]:
    """Recupera i valori tradotti dal template senza rieseguire filtri linguistici."""
    pattern_parts: list[str] = []
    position = 0
    for match in PLACEHOLDER_RE.finditer(template):
        pattern_parts.append(re.escape(template[position : match.start()]))
        pattern_parts.append(f"(?P<{match.group(1)}>.*?)")
        position = match.end()
    pattern_parts.append(re.escape(template[position:]))
    matched = re.fullmatch("".join(pattern_parts), rendered, flags=re.DOTALL)
    if matched is None:
        raise ValueError("Lo staging esistente non conserva il template strutturale")
    recovered = {identifier: matched.group(identifier) for identifier in source}
    if not all(value.strip() for value in recovered.values()):
        raise ValueError("Lo staging esistente contiene traduzioni vuote")
    return recovered


def derive_disclaimer_txt(
    source_html: str, localized_html: str, source_txt: str,
    html_lang: str, direction: str,
) -> str:
    """Deriva il disclaimer TXT dall'HTML localizzato e revisionato."""
    html_template, html_source_units = extract_html_text(source_html)
    localized_template = update_html_language(html_template, html_lang, direction)
    html_target_units = recover_translations(
        localized_template, html_source_units, localized_html,
    )
    pairs = [
        (
            html_module.unescape(html_source_units[identifier]).strip(),
            html_module.unescape(html_target_units[identifier]).strip(),
        )
        for identifier in html_source_units
    ]

    output: list[str] = []
    for line in source_txt.splitlines(keepends=True):
        ending = "\r\n" if line.endswith("\r\n") else "\n" if line.endswith("\n") else ""
        value = line[: -len(ending)] if ending else line
        if not should_translate(value):
            output.append(line)
            continue
        bullet = "- " if value.startswith("- ") else ""
        body = value[len(bullet) :]
        normalized_body = html_module.unescape(body).strip()
        candidates = [
            (source_value, target_value)
            for source_value, target_value in pairs
            if normalized_body.casefold().startswith(source_value.casefold())
        ]
        if not candidates:
            raise ValueError(f"Riga TXT non derivabile dal disclaimer HTML: {value}")
        source_value, target_value = max(candidates, key=lambda pair: len(pair[0]))
        suffix = normalized_body[len(source_value) :]
        translated = target_value + suffix
        if normalized_body == normalized_body.upper() and any(char.isalpha() for char in normalized_body):
            translated = translated.upper()
        output.append(f"{bullet}{translated}{ending}")
    return "".join(output)


def update_html_language(html: str, html_lang: str, direction: str) -> str:
    match = re.search(r"<html\b([^>]*)>", html, flags=re.IGNORECASE)
    if not match:
        raise ValueError("Elemento <html> non trovato")
    attributes = match.group(1)
    if re.search(r"\blang=[\"\'][^\"\']*[\"\']", attributes, flags=re.I):
        attributes = re.sub(
            r"\blang=[\"\'][^\"\']*[\"\']", f'lang="{html_lang}"', attributes,
            count=1, flags=re.I,
        )
    else:
        attributes += f' lang="{html_lang}"'
    if re.search(r"\bdir=[\"\'][^\"\']*[\"\']", attributes, flags=re.I):
        attributes = re.sub(
            r"\bdir=[\"\'][^\"\']*[\"\']", f'dir="{direction}"', attributes,
            count=1, flags=re.I,
        )
    else:
        attributes += f' dir="{direction}"'
    return html[: match.start()] + f"<html{attributes}>" + html[match.end() :]


def load_api_key() -> str:
    key = os.environ.get("ATK_PRO_GEMINI_API_KEY", "").strip()
    if key:
        return key
    if CONFIG_FILE.is_file():
        data = json.loads(CONFIG_FILE.read_text(encoding="utf-8"))
        key = str(data.get("ocr_api_key", "")).strip()
    if not key:
        raise RuntimeError(
            "Chiave Gemini assente: impostare ATK_PRO_GEMINI_API_KEY o ocr_api_key in config.json"
        )
    return key


def load_vault_keys(keys_file: Path) -> dict[str, str]:
    """Carica una sola chiave per provider senza mai stamparne il valore."""
    if not keys_file.is_file():
        raise FileNotFoundError(f"Caveau chiavi non trovato: {keys_file}")
    lines = keys_file.read_text(encoding="utf-8-sig").splitlines()
    delimiter = ";"
    if lines and lines[0].lower().startswith("sep="):
        delimiter = lines[0][4:5] or ";"
        lines = lines[1:]
    keys: dict[str, str] = {}
    for row in csv.DictReader(lines, delimiter=delimiter):
        provider = str(row.get("Provider") or "").strip()
        key = str(row.get("Key") or "").strip()
        if provider and key and provider not in keys:
            keys[provider] = key
    return keys


def post_json(
    url: str, headers: dict[str, str], payload: dict[str, Any], timeout: int = 240,
) -> dict[str, Any]:
    request = urllib.request.Request(
        url,
        data=json.dumps(payload, ensure_ascii=False).encode("utf-8"),
        headers={"Content-Type": "application/json", **headers},
        method="POST",
    )
    last_error: Exception | None = None
    for attempt in range(3):
        try:
            with urllib.request.urlopen(request, timeout=timeout) as response:
                return json.load(response)
        except urllib.error.HTTPError as error:
            detail = error.read().decode("utf-8", errors="replace")[:800]
            last_error = RuntimeError(f"HTTP {error.code}: {detail}")
            if error.code not in (429, 500, 502, 503, 504):
                raise last_error
        except Exception as error:
            last_error = error
        time.sleep(5 * (attempt + 1))
    raise RuntimeError(str(last_error))


def parse_json_response(raw: str) -> dict[str, str]:
    raw = raw.strip()
    raw = re.sub(r"^```(?:json)?\s*", "", raw)
    raw = re.sub(r"\s*```$", "", raw)
    start, end = raw.find("{"), raw.rfind("}")
    if start < 0 or end < start:
        raise ValueError("Risposta JSON non riconoscibile")
    parsed = json.loads(raw[start : end + 1])
    if not isinstance(parsed, dict):
        raise ValueError("La risposta del provider non è un oggetto JSON")
    return parsed


def build_translation_prompt(texts: dict[str, str], language_name: str) -> str:
    return (
        f"Translate every JSON value from Italian to {language_name}. Return only one valid JSON "
        "object with exactly the same keys. Preserve the complete meaning, including legal scope, "
        "prohibitions, limitations, qualifications and obligations. Use natural publication-quality "
        "language. Preserve HTML entities, symbols, numbers, URLs and file names. Never translate "
        f"these expressions when present: {', '.join(DO_NOT_TRANSLATE)}. Do not add notes, caveats "
        "or statements that another language prevails.\n\n"
        + json.dumps(texts, ensure_ascii=False, indent=2)
    )


def chunks(texts: dict[str, str], size: int = 40) -> list[dict[str, str]]:
    keys = list(texts)
    return [
        {key: texts[key] for key in keys[offset : offset + size]}
        for offset in range(0, len(keys), size)
    ]


def load_translation_memory(path: Path | None, language: str) -> dict[str, str]:
    """Legge le corrispondenze esatte IT -> lingua dal workbook storico."""
    if path is None:
        return {}
    if not path.is_file():
        raise FileNotFoundError(f"Memoria di traduzione non trovata: {path}")
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook["Translations"]
    headers = [str(cell.value or "") for cell in sheet[1]]
    try:
        italian_column = headers.index("Italiano") + 1
        target_column = headers.index(MEMORY_HEADERS[language]) + 1
    except (KeyError, ValueError) as error:
        raise ValueError(f"Colonna memoria assente per {language}") from error
    memory: dict[str, str] = {}
    for row in range(2, sheet.max_row + 1):
        source = sheet.cell(row, italian_column).value
        target = sheet.cell(row, target_column).value
        if source and target:
            memory.setdefault(str(source), str(target))
    workbook.close()
    return memory


def reference_for_ids(
    texts: dict[str, str], translation_memory: dict[str, str],
) -> dict[str, str]:
    return {
        identifier: translation_memory[source]
        for identifier, source in texts.items()
        if source in translation_memory
    }


def with_reference(prompt: str, references: dict[str, str] | None) -> str:
    if not references:
        return prompt
    return (
        prompt
        + "\n\nHISTORICAL REFERENCE TRANSLATIONS:\n"
        + json.dumps(references, ensure_ascii=False, indent=2)
        + "\nUse these as terminology references only. Correct them when they are outdated, "
        "unnatural, incomplete, or inconsistent with the current Italian source."
    )


class GeminiTranslator:
    def __init__(self, model_name: str, delay_seconds: float) -> None:
        import google.generativeai as genai

        genai.configure(api_key=load_api_key())
        self.model = genai.GenerativeModel(model_name)
        self.delay_seconds = delay_seconds

    def translate(
        self, texts: dict[str, str], language_name: str,
        references: dict[str, str] | None = None,
    ) -> dict[str, str]:
        translated: dict[str, str] = {}
        keys = list(texts)
        for offset in range(0, len(keys), 40):
            chunk_keys = keys[offset : offset + 40]
            chunk = {key: texts[key] for key in chunk_keys}
            prompt = with_reference((
                f"Translate every JSON value from Italian to {language_name}.\n"
                "Return only one valid JSON object with exactly the same keys.\n"
                "Preserve meaning, paragraph function, punctuation, URLs, file names, "
                "HTML entities, symbols and legal precision.\n"
                f"Never translate these expressions: {', '.join(DO_NOT_TRANSLATE)}.\n"
                "Do not add qualifications such as informal translation or precedence "
                "of another language.\n\n"
                + json.dumps(chunk, ensure_ascii=False, indent=2)
            ), {key: references[key] for key in chunk if references and key in references})
            response = self.model.generate_content(prompt)
            raw = getattr(response, "text", "").strip()
            raw = re.sub(r"^```(?:json)?\s*", "", raw)
            raw = re.sub(r"\s*```$", "", raw)
            result = json.loads(raw)
            if set(result) != set(chunk):
                raise ValueError("La risposta del provider non conserva i segnaposto del blocco")
            translated.update(result)
            if offset + 40 < len(keys):
                time.sleep(self.delay_seconds)
        return translated


class OpenAITranslator:
    def __init__(self, api_key: str, model_name: str, delay_seconds: float) -> None:
        self.api_key = api_key
        self.model_name = model_name
        self.delay_seconds = delay_seconds

    def translate(
        self, texts: dict[str, str], language_name: str,
        references: dict[str, str] | None = None,
    ) -> dict[str, str]:
        translated: dict[str, str] = {}
        blocks = chunks(texts)
        for index, block in enumerate(blocks):
            data = post_json(
                "https://api.openai.com/v1/responses",
                {"Authorization": f"Bearer {self.api_key}"},
                {
                    "model": self.model_name,
                    "input": with_reference(
                        build_translation_prompt(block, language_name),
                        {key: references[key] for key in block if references and key in references},
                    ),
                    "max_output_tokens": 12000,
                },
            )
            raw = "".join(
                content.get("text", "")
                for item in data.get("output", [])
                for content in item.get("content", [])
                if content.get("type") == "output_text"
            )
            result = parse_json_response(raw)
            if set(result) != set(block):
                raise ValueError("OpenAI non ha conservato tutti i segnaposto")
            translated.update(result)
            if index + 1 < len(blocks):
                time.sleep(self.delay_seconds)
        return translated


class ClaudeReviewer:
    def __init__(self, api_key: str, model_name: str, delay_seconds: float) -> None:
        self.api_key = api_key
        self.model_name = model_name
        self.delay_seconds = delay_seconds

    def review(
        self, source: dict[str, str], draft: dict[str, str], language_name: str,
        references: dict[str, str] | None = None,
    ) -> dict[str, str]:
        reviewed: dict[str, str] = {}
        source_blocks = chunks(source)
        for index, source_block in enumerate(source_blocks):
            draft_block = {key: draft[key] for key in source_block}
            prompt = with_reference((
                f"Review the draft translations from Italian to {language_name}. Correct every value "
                "that is incomplete, unnatural, terminologically imprecise or changes legal scope. "
                "Preserve prohibitions, limitations, qualifications and obligations exactly. Preserve "
                "numbers, URLs, file names, HTML entities and invariant product or licence names. Return "
                "only one JSON object with exactly the same keys and the final corrected translations. "
                "Do not add notes or commentary.\n\nSOURCE:\n"
                + json.dumps(source_block, ensure_ascii=False, indent=2)
                + "\n\nDRAFT:\n"
                + json.dumps(draft_block, ensure_ascii=False, indent=2)
            ), {key: references[key] for key in source_block if references and key in references})
            data = post_json(
                "https://api.anthropic.com/v1/messages",
                {"x-api-key": self.api_key, "anthropic-version": "2023-06-01"},
                {
                    "model": self.model_name,
                    "max_tokens": 12000,
                    "messages": [{"role": "user", "content": prompt}],
                },
            )
            raw = "".join(block.get("text", "") for block in data.get("content", []))
            result = parse_json_response(raw)
            if set(result) != set(source_block):
                raise ValueError("Claude non ha conservato tutti i segnaposto")
            reviewed.update(result)
            if index + 1 < len(source_blocks):
                time.sleep(self.delay_seconds)
        return reviewed

    def resolve_conflicts(
        self, conflicts: dict[str, list[str]], language_name: str,
        historical: dict[str, str],
    ) -> dict[str, str]:
        """Sceglie una formulazione canonica per ogni sorgente italiana ripetuta."""
        identifiers = {f"C{index:04d}": source for index, source in enumerate(conflicts)}
        choices = {
            identifier: {
                "italian": source,
                "current_variants": conflicts[source],
                "historical_reference": historical.get(source),
            }
            for identifier, source in identifiers.items()
        }
        prompt = (
            f"For each item, select or produce one canonical {language_name} translation of the "
            "Italian source. It must work consistently everywhere the exact same Italian text occurs. "
            "Correct incomplete fragments, preserve legal scope and keep product, licence, file and "
            "technical names unchanged. Historical references are advisory only. Return only a JSON "
            "object mapping every C-key to the canonical translated string.\n\n"
            + json.dumps(choices, ensure_ascii=False, indent=2)
        )
        data = post_json(
            "https://api.anthropic.com/v1/messages",
            {"x-api-key": self.api_key, "anthropic-version": "2023-06-01"},
            {
                "model": self.model_name,
                "max_tokens": 12000,
                "messages": [{"role": "user", "content": prompt}],
            },
        )
        raw = "".join(block.get("text", "") for block in data.get("content", []))
        resolved = parse_json_response(raw)
        if set(resolved) != set(identifiers):
            raise ValueError("Claude non ha restituito tutte le scelte canoniche")
        return {source: resolved[identifier] for identifier, source in identifiers.items()}


def translate_using_shared_memory(
    texts: dict[str, str], language_name: str, translator: Any,
    shared: dict[str, str], historical: dict[str, str],
) -> dict[str, str]:
    """Traduce una sola volta ogni sorgente identica e riusa la forma canonica."""
    pending_by_source: dict[str, str] = {}
    result: dict[str, str] = {}
    for identifier, source in texts.items():
        if source in shared:
            result[identifier] = shared[source]
        else:
            pending_by_source.setdefault(source, identifier)
    pending = {identifier: source for source, identifier in pending_by_source.items()}
    if pending:
        references = reference_for_ids(pending, historical)
        translated = translator.translate(pending, language_name, references)
        for identifier, source in pending.items():
            shared[source] = translated[identifier]
    for identifier, source in texts.items():
        result[identifier] = shared[source]
    return result


def review_using_shared_memory(
    source: dict[str, str], draft: dict[str, str], language_name: str,
    reviewer: ClaudeReviewer, shared: dict[str, str], historical: dict[str, str],
) -> dict[str, str]:
    pending_by_source: dict[str, str] = {}
    result: dict[str, str] = {}
    for identifier, source_text in source.items():
        if source_text in shared:
            result[identifier] = shared[source_text]
        else:
            pending_by_source.setdefault(source_text, identifier)
    pending_source = {identifier: text for text, identifier in pending_by_source.items()}
    if pending_source:
        pending_draft = {identifier: draft[identifier] for identifier in pending_source}
        references = reference_for_ids(pending_source, historical)
        reviewed = reviewer.review(
            pending_source, pending_draft, language_name, references,
        )
        for identifier, source_text in pending_source.items():
            shared[source_text] = reviewed[identifier]
    for identifier, source_text in source.items():
        result[identifier] = shared[source_text]
    return result


def audit_sources() -> dict[str, dict[str, Any]]:
    report: dict[str, dict[str, Any]] = {}
    missing = [name for name in DOCUMENT_FILES if not (ITALIAN_DIR / name).is_file()]
    if missing:
        raise FileNotFoundError(f"Sorgenti italiane mancanti: {', '.join(missing)}")

    expected_links = set(GUIDE_FILES[1:])
    guide_html = (ITALIAN_DIR / "guida.html").read_text(encoding="utf-8")
    actual_links = {Path(link).name for link in LINK_RE.findall(guide_html)}
    absent_links = sorted(expected_links - actual_links)
    if absent_links:
        raise ValueError(f"Collegamenti mancanti dalla guida italiana: {absent_links}")

    for name in DOCUMENT_FILES:
        source = (ITALIAN_DIR / name).read_text(encoding="utf-8")
        _, texts = extract_html_text(source) if name.endswith(".html") else extract_plain_text(source)
        report[name] = {
            "sha256": sha256_text(source),
            "characters": len(source),
            "translation_units": len(texts),
        }
    return report


def resolve_languages(only: list[str] | None, from_language: str | None) -> list[str]:
    selected = list(LANGUAGES)
    if only:
        unknown = sorted(set(only) - set(LANGUAGES))
        if unknown:
            raise ValueError(f"Lingue non supportate: {', '.join(unknown)}")
        selected = [code for code in selected if code in only]
    if from_language:
        if from_language not in selected:
            raise ValueError("--from-lang deve appartenere alle lingue selezionate")
        selected = selected[selected.index(from_language) :]
    return selected


def safe_destination(staging_root: Path, language: str, filename: str) -> Path:
    root = staging_root.resolve()
    destination = (root / language / "testuali" / filename).resolve()
    if root == ASSETS_DIR.resolve() or ASSETS_DIR.resolve() in destination.parents:
        raise ValueError("La destinazione non può trovarsi dentro assets")
    if language == "it":
        raise ValueError("La sorgente italiana è protetta da scrittura")
    if root not in destination.parents:
        raise ValueError("Destinazione esterna all'area di revisione")
    return destination


def execute_translation(
    languages: list[str], staging_root: Path, provider: str, model_name: str,
    review_provider: str | None, review_model: str, keys_file: Path | None,
    translation_memory_path: Path | None, delay_seconds: float, overwrite: bool,
    resume: bool,
) -> None:
    vault_keys = load_vault_keys(keys_file) if keys_file else {}
    if provider == "openai":
        if "OpenAI" not in vault_keys:
            raise RuntimeError("Chiave OpenAI assente dal caveau")
        translator: Any = OpenAITranslator(vault_keys["OpenAI"], model_name, delay_seconds)
    elif provider == "gemini":
        translator = GeminiTranslator(model_name, delay_seconds)
    else:
        raise ValueError(f"Provider non supportato: {provider}")

    reviewer: ClaudeReviewer | None = None
    if review_provider == "claude":
        if "Claude" not in vault_keys:
            raise RuntimeError("Chiave Claude assente dal caveau")
        reviewer = ClaudeReviewer(vault_keys["Claude"], review_model, delay_seconds)

    draft_root = staging_root / f"draft-{provider}"
    reviewed_root = staging_root / f"reviewed-{review_provider}" if reviewer else None
    manifest: dict[str, Any] = {
        "source_language": "it",
        "provider": provider,
        "model": model_name,
        "review_provider": review_provider,
        "review_model": review_model if reviewer else None,
        "documents": {},
    }
    for language in languages:
        language_name, html_lang, direction = LANGUAGES[language]
        historical = load_translation_memory(translation_memory_path, language)
        shared_draft: dict[str, str] = {}
        shared_reviewed: dict[str, str] = {}
        for filename in DOCUMENT_FILES:
            source_path = ITALIAN_DIR / filename
            draft_destination = safe_destination(draft_root, language, filename)
            reviewed_destination = (
                safe_destination(reviewed_root, language, filename) if reviewed_root else None
            )
            existing = [path for path in (draft_destination, reviewed_destination) if path and path.exists()]
            if existing and not overwrite and not resume:
                raise FileExistsError(
                    f"File di staging già presente: {existing[0]}; usare --overwrite-staging"
                )
            source = source_path.read_text(encoding="utf-8")
            template, texts = (
                extract_html_text(source) if filename.endswith(".html") else extract_plain_text(source)
            )
            rendered_template = (
                update_html_language(template, html_lang, direction)
                if filename.endswith(".html") else template
            )
            if filename == "disclaimer_legale_ATK-Pro.txt":
                source_html = (ITALIAN_DIR / "disclaimer_legale_ATK-Pro.html").read_text(
                    encoding="utf-8"
                )
                draft_html_path = safe_destination(
                    draft_root, language, "disclaimer_legale_ATK-Pro.html"
                )
                if not draft_html_path.is_file():
                    raise ValueError("Disclaimer HTML draft assente: impossibile derivare il TXT")
                draft_rendered = derive_disclaimer_txt(
                    source_html,
                    draft_html_path.read_text(encoding="utf-8"),
                    source,
                    html_lang,
                    direction,
                )
                draft_destination.parent.mkdir(parents=True, exist_ok=True)
                draft_destination.write_text(draft_rendered, encoding="utf-8", newline="")
                final_rendered = draft_rendered
                if reviewed_destination:
                    reviewed_html_path = safe_destination(
                        reviewed_root, language, "disclaimer_legale_ATK-Pro.html"
                    )
                    if not reviewed_html_path.is_file():
                        raise ValueError("Disclaimer HTML revisionato assente: impossibile derivare il TXT")
                    final_rendered = derive_disclaimer_txt(
                        source_html,
                        reviewed_html_path.read_text(encoding="utf-8"),
                        source,
                        html_lang,
                        direction,
                    )
                    reviewed_destination.parent.mkdir(parents=True, exist_ok=True)
                    reviewed_destination.write_text(final_rendered, encoding="utf-8", newline="")
                manifest["documents"][f"{language}/{filename}"] = {
                    "source_sha256": sha256_text(source),
                    "draft_sha256": sha256_text(draft_rendered),
                    "output_sha256": sha256_text(final_rendered),
                    "translation_units": len(texts),
                    "historical_matches": sum(1 for text in texts.values() if text in historical),
                    "derived_from": "disclaimer_legale_ATK-Pro.html",
                }
                print(f"DERIVE {language}/{filename}: {len(texts)} unità dall'HTML")
                continue
            complete_pair = draft_destination.is_file() and (
                reviewed_destination is None or reviewed_destination.is_file()
            )
            if resume and complete_pair:
                draft_rendered = draft_destination.read_text(encoding="utf-8")
                final_rendered = (
                    reviewed_destination.read_text(encoding="utf-8")
                    if reviewed_destination else draft_rendered
                )
                try:
                    draft_units = recover_translations(rendered_template, texts, draft_rendered)
                    final_units = recover_translations(rendered_template, texts, final_rendered)
                except ValueError as error:
                    raise ValueError(
                        f"Staging esistente non allineato per {language}/{filename}: {error}"
                    ) from error
                for identifier, source_text in texts.items():
                    shared_draft.setdefault(source_text, draft_units[identifier])
                    shared_reviewed.setdefault(source_text, final_units[identifier])
                manifest["documents"][f"{language}/{filename}"] = {
                    "source_sha256": sha256_text(source),
                    "draft_sha256": sha256_text(draft_rendered),
                    "output_sha256": sha256_text(final_rendered),
                    "translation_units": len(texts),
                    "historical_matches": sum(1 for text in texts.values() if text in historical),
                }
                print(f"SKIP {language}/{filename}: staging completo e verificato")
                continue
            reuse_complete_draft = resume and draft_destination.is_file()
            if reuse_complete_draft:
                draft_rendered = draft_destination.read_text(encoding="utf-8")
                try:
                    translated = recover_translations(rendered_template, texts, draft_rendered)
                except ValueError as error:
                    raise ValueError(
                        f"Draft esistente non allineato per {language}/{filename}: {error}"
                    ) from error
                for identifier, source_text in texts.items():
                    shared_draft.setdefault(source_text, translated[identifier])
                print(f"REUSE {language}/{filename}: draft completo e verificato")
            else:
                translated = translate_using_shared_memory(
                    texts, language_name, translator, shared_draft, historical,
                )
                draft_rendered = apply_translations(template, texts, translated)
                if filename.endswith(".html"):
                    draft_rendered = update_html_language(draft_rendered, html_lang, direction)
                draft_destination.parent.mkdir(parents=True, exist_ok=True)
                draft_destination.write_text(draft_rendered, encoding="utf-8", newline="")

            final_translations = (
                review_using_shared_memory(
                    texts, translated, language_name, reviewer,
                    shared_reviewed, historical,
                )
                if reviewer else translated
            )
            final_rendered = apply_translations(template, texts, final_translations)
            if filename.endswith(".html"):
                final_rendered = update_html_language(final_rendered, html_lang, direction)
            if reviewed_destination:
                reviewed_destination.parent.mkdir(parents=True, exist_ok=True)
                reviewed_destination.write_text(final_rendered, encoding="utf-8", newline="")
            manifest["documents"][f"{language}/{filename}"] = {
                "source_sha256": sha256_text(source),
                "draft_sha256": sha256_text(draft_rendered),
                "output_sha256": sha256_text(final_rendered),
                "translation_units": len(texts),
                "historical_matches": sum(1 for text in texts.values() if text in historical),
            }
            print(f"OK  {language}/{filename}: {len(texts)} unità")

    manifest_path = staging_root.resolve() / "translation_manifest.json"
    manifest_path.write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    print(f"Manifest: {manifest_path}")


def reconcile_existing_staging(
    languages: list[str], staging_root: Path, review_provider: str | None,
    review_model: str, keys_file: Path | None,
    translation_memory_path: Path | None,
) -> None:
    if review_provider != "claude" or keys_file is None:
        raise ValueError("La riconciliazione richiede --review-provider claude e --keys-file")
    vault_keys = load_vault_keys(keys_file)
    if "Claude" not in vault_keys:
        raise RuntimeError("Chiave Claude assente dal caveau")
    reviewer = ClaudeReviewer(vault_keys["Claude"], review_model, 0)

    for language in languages:
        language_name = LANGUAGES[language][0]
        historical = load_translation_memory(translation_memory_path, language)
        target_root = staging_root / "reviewed-claude" / language / "testuali"
        documents: dict[str, tuple[str, dict[str, str], list[str]]] = {}
        variants: dict[str, list[str]] = {}

        for filename in DOCUMENT_FILES:
            source_text = (ITALIAN_DIR / filename).read_text(encoding="utf-8")
            target_path = target_root / filename
            if not target_path.is_file():
                raise FileNotFoundError(f"Staging revisionato mancante: {target_path}")
            target_text = target_path.read_text(encoding="utf-8")
            extractor = extract_html_text if filename.endswith(".html") else extract_plain_text
            _, source_units = extractor(source_text)
            target_template, target_units = extractor(target_text)
            if len(source_units) != len(target_units):
                raise ValueError(f"Unità non allineate in {filename}")
            source_values = list(source_units.values())
            for source_value, target_value in zip(source_values, target_units.values()):
                variants.setdefault(source_value, [])
                if target_value not in variants[source_value]:
                    variants[source_value].append(target_value)
            documents[filename] = (target_template, target_units, source_values)

        conflicts = {source: values for source, values in variants.items() if len(values) > 1}
        resolved = reviewer.resolve_conflicts(conflicts, language_name, historical) if conflicts else {}
        canonical = {
            source: resolved.get(source, values[0])
            for source, values in variants.items()
        }

        for filename, (target_template, target_units, source_values) in documents.items():
            replacements = {
                identifier: canonical[source_value]
                for identifier, source_value in zip(target_units, source_values)
            }
            rendered = apply_translations(target_template, target_units, replacements)
            (target_root / filename).write_text(rendered, encoding="utf-8", newline="")

        report = {
            "language": language,
            "review_provider": review_provider,
            "review_model": review_model,
            "historical_memory_entries": len(historical),
            "conflicts_before": len(conflicts),
            "conflicts_after": 0,
            "choices": {
                source: {
                    "previous_variants": conflicts[source],
                    "historical_reference": historical.get(source),
                    "canonical": resolved[source],
                }
                for source in conflicts
            },
        }
        report_path = staging_root / f"reconciliation_report_{language}.json"
        report_path.write_text(
            json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
        )
        print(f"Riconciliazione {language}: {len(conflicts)} conflitti risolti")
        print(f"Rapporto: {report_path}")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Audita o prepara in staging i 14 documenti localizzati di ATK-Pro"
    )
    parser.add_argument("--only", nargs="+", help="Limita le lingue, es. --only en")
    parser.add_argument("--from-lang", help="Riprende dalla lingua indicata")
    parser.add_argument(
        "--execute", action="store_true",
        help="Autorizza chiamate esterne e scrittura esclusivamente nello staging",
    )
    parser.add_argument(
        "--reconcile-existing", action="store_true",
        help="Uniforma le traduzioni ripetute nello staging revisionato esistente",
    )
    parser.add_argument("--provider", choices=("openai", "gemini"), default="openai")
    parser.add_argument("--model", default="gpt-5.6-sol")
    parser.add_argument("--review-provider", choices=("claude",))
    parser.add_argument("--review-model", default="claude-sonnet-5")
    parser.add_argument("--keys-file", type=Path)
    parser.add_argument(
        "--translation-memory", type=Path, default=DEFAULT_TRANSLATION_MEMORY,
        help="Workbook storico usato come riferimento terminologico non vincolante",
    )
    parser.add_argument("--delay-seconds", type=float, default=10.0)
    parser.add_argument("--staging-root", type=Path, default=DEFAULT_STAGING_ROOT)
    parser.add_argument("--overwrite-staging", action="store_true")
    parser.add_argument(
        "--resume-staging", action="store_true",
        help="Riprende uno staging incompleto e salta solo le coppie già complete e strutturalmente valide",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    try:
        languages = resolve_languages(args.only, args.from_lang)
        report = audit_sources()
        units = sum(item["translation_units"] for item in report.values())
        print("Audit localizzazione documentale superato.")
        print(f"Sorgenti italiane: {len(DOCUMENT_FILES)} (guide 10, presentazioni 2, disclaimer 2)")
        print(f"Lingue selezionate: {len(languages)} ({', '.join(languages)})")
        print(f"Unità testuali per lingua: {units}")
        print(f"Output eventuali: {len(DOCUMENT_FILES) * len(languages)}")
        for filename, details in report.items():
            print(f"- {filename}: {details['translation_units']} unità")
        if args.reconcile_existing:
            reconcile_existing_staging(
                languages, args.staging_root, args.review_provider,
                args.review_model, args.keys_file, args.translation_memory,
            )
            return 0
        if not args.execute:
            print("Modalità audit: nessuna API chiamata e nessun file scritto.")
            return 0
        execute_translation(
            languages, args.staging_root, args.provider, args.model,
            args.review_provider, args.review_model, args.keys_file,
            args.translation_memory, args.delay_seconds, args.overwrite_staging,
            args.resume_staging,
        )
        return 0
    except Exception as error:
        print(f"ERRORE: {error}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
